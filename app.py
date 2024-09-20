from flask import Flask, request, jsonify
import openpyxl
import os
from datetime import datetime

app = Flask(__name__)

def load_or_create_excel(file_name):
    if os.path.exists(file_name):
        workbook = openpyxl.load_workbook(file_name)
    else:
        workbook = openpyxl.Workbook()
        workbook.active.append(["Timestamp", "Ph_Value", "Turbidity", "Temperature", "Flow_Value"])  # Header
        workbook.save(file_name)
    return workbook

# API route to handle POST requests and save data to Excel
@app.route('/add_to_excel', methods=['POST'])
def add_to_excel():
    data = request.get_json()

    # Automatically generate the timestamp
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    ph_value = data.get('ph_value')
    turbidity = data.get('turbidity')
    temperature = data.get('temperature')
    flow_value = data.get('flow_value')

    if not all([ph_value, turbidity, temperature, flow_value]):
        return jsonify({"error": "Missing required parameters"}), 400

    # File name for the Excel file
    file_name = "data.xlsx"
    workbook = load_or_create_excel(file_name)
    sheet = workbook.active

    # Add the data to the Excel sheet
    sheet.append([timestamp, ph_value, turbidity, temperature, flow_value])
    workbook.save(file_name)

    return jsonify({"message": "Data added to Excel successfully!"}), 200

# API route to retrieve the latest row from the Excel file
@app.route('/get_entry', methods=['GET'])
def get_latest_entry():
    # File name for the Excel file
    file_name = "data.xlsx"
    
    if not os.path.exists(file_name):
        return jsonify({"error": "Excel file not found"}), 404

    workbook = openpyxl.load_workbook(file_name)
    sheet = workbook.active

    if sheet.max_row < 2:  # Check if there are any rows other than the header
        return jsonify({"error": "No data available"}), 404

    # Get the latest row (excluding the header)
    latest_row = list(sheet.iter_rows(values_only=True))[-1]
    
    # Create an object for the latest entry
    latest_entry = {
        "timestamp": latest_row[0],
        "ph_value": latest_row[1],
        "turbidity": latest_row[2],
        "temperature": latest_row[3],
        "flow_value": latest_row[4]
    }

    return jsonify({"latest_entry": latest_entry}), 200

if __name__ == '__main__':
    app.run(debug=True)
